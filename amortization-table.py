#!/usr/bin/env python3
"""
Amortization table generator (fixed-rate, fully amortizing loan) with prompts.

Outputs:
- amortization_schedule.csv
- amortization_schedule.xlsx (if openpyxl is installed)

Prompts for:
- principal
- annual interest rate (percent)
- term (months)
- optional extra principal per payment
"""

from __future__ import annotations

from dataclasses import dataclass
from decimal import Decimal, ROUND_HALF_UP, getcontext
from pathlib import Path
from typing import List, Dict
import csv

try:
    import openpyxl
except ImportError:
    openpyxl = None

getcontext().prec = 28
CENT = Decimal("0.01")


def money(x: Decimal) -> Decimal:
    return x.quantize(CENT, rounding=ROUND_HALF_UP)


def parse_decimal(prompt: str, min_value: Decimal | None = None) -> Decimal:
    while True:
        raw = input(prompt).strip().replace(",", "")
        try:
            val = Decimal(raw)
            if min_value is not None and val < min_value:
                print(f"Value must be >= {min_value}. Try again.")
                continue
            return val
        except Exception:
            print("Invalid number. Try again (example: 100000 or 100000.00).")


def parse_int(prompt: str, min_value: int | None = None) -> int:
    while True:
        raw = input(prompt).strip().replace(",", "")
        try:
            val = int(raw)
            if min_value is not None and val < min_value:
                print(f"Value must be >= {min_value}. Try again.")
                continue
            return val
        except Exception:
            print("Invalid integer. Try again (example: 120).")


def ask_yes_no(prompt: str, default_yes: bool = True) -> bool:
    default = "Y/n" if default_yes else "y/N"
    while True:
        raw = input(f"{prompt} ({default}): ").strip().lower()
        if raw == "":
            return default_yes
        if raw in ("y", "yes"):
            return True
        if raw in ("n", "no"):
            return False
        print("Please enter y or n.")


@dataclass
class LoanInputs:
    principal: Decimal          # e.g. 100000.00
    annual_rate: Decimal        # e.g. 0.03 for 3%
    term_months: int            # e.g. 120
    extra_principal: Decimal    # e.g. 50.00


def calc_monthly_payment(principal: Decimal, annual_rate: Decimal, term_months: int) -> Decimal:
    if term_months <= 0:
        raise ValueError("term_months must be > 0")

    monthly_rate = annual_rate / Decimal("12")
    if monthly_rate == 0:
        return money(principal / Decimal(term_months))

    r = monthly_rate
    n = Decimal(term_months)
    payment = (r * principal) / (Decimal("1") - (Decimal("1") + r) ** (-n))
    return money(payment)


def build_schedule(inputs: LoanInputs) -> List[Dict[str, object]]:
    balance = inputs.principal
    monthly_rate = inputs.annual_rate / Decimal("12")
    base_payment = calc_monthly_payment(inputs.principal, inputs.annual_rate, inputs.term_months)

    rows: List[Dict[str, object]] = []

    for period in range(1, inputs.term_months + 1):
        if balance <= 0:
            break

        interest = money(balance * monthly_rate)
        scheduled_principal = base_payment - interest
        if scheduled_principal < 0:
            raise ValueError("Payment is too small to cover interest. Check rate/term.")

        extra = inputs.extra_principal
        total_principal = scheduled_principal + extra

        # Prevent overpay on final period
        if total_principal > balance:
            total_principal = balance
            extra = money(total_principal - scheduled_principal) if total_principal >= scheduled_principal else Decimal("0.00")
            payment = money(interest + total_principal)
        else:
            payment = base_payment

        new_balance = money(balance - total_principal)

        rows.append(
            {
                "Period": period,
                "Payment": float(payment),
                "Interest": float(interest),
                "Principal": float(money(scheduled_principal)),
                "ExtraPrincipal": float(money(extra)),
                "TotalPrincipal": float(money(total_principal)),
                "Balance": float(new_balance),
            }
        )
        balance = new_balance

    return rows


def write_csv(rows: List[Dict[str, object]], out_path: Path) -> None:
    headers = list(rows[0].keys()) if rows else []
    with out_path.open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=headers)
        w.writeheader()
        for r in rows:
            w.writerow(r)


def write_xlsx(rows: List[Dict[str, object]], out_path: Path) -> None:
    if openpyxl is None:
        return

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Amortization"

    if not rows:
        wb.save(out_path)
        return

    headers = list(rows[0].keys())
    ws.append(headers)
    for r in rows:
        ws.append([r[h] for h in headers])

    # Formatting
    currency_cols = {"Payment", "Interest", "Principal", "ExtraPrincipal", "TotalPrincipal", "Balance"}
    for col_idx, header in enumerate(headers, start=1):
        if header in currency_cols:
            for row_idx in range(2, ws.max_row + 1):
                ws.cell(row=row_idx, column=col_idx).number_format = '"$"#,##0.00'

    ws.freeze_panes = "A2"
    wb.save(out_path)


def prompt_inputs() -> LoanInputs:
    print("\n--- Amortization Table Inputs ---")
    principal = parse_decimal("Loan principal (e.g., 100000.00): ", min_value=Decimal("0.01"))
    rate_pct = parse_decimal("Annual interest rate % (e.g., 3 for 3%): ", min_value=Decimal("0.00"))
    term_months = parse_int("Term in months (e.g., 120): ", min_value=1)
    extra = parse_decimal("Extra principal per payment (optional, default 0): ", min_value=Decimal("0.00"))

    return LoanInputs(
        principal=principal,
        annual_rate=rate_pct / Decimal("100"),
        term_months=term_months,
        extra_principal=extra,
    )


def main() -> None:
    inputs = prompt_inputs()

    rows = build_schedule(inputs)

    out_dir = Path(".")
    csv_path = out_dir / "amortization_schedule.csv"
    xlsx_path = out_dir / "amortization_schedule.xlsx"

    write_csv(rows, csv_path)
    print(f"\nCreated: {csv_path.resolve()}")

    if openpyxl is None:
        if ask_yes_no("openpyxl not installed. Skip Excel output?", default_yes=True):
            pass
        else:
            print("Install it with: pip install openpyxl")
    else:
        write_xlsx(rows, xlsx_path)
        print(f"Created: {xlsx_path.resolve()}")

    total_interest = sum(Decimal(str(r["Interest"])) for r in rows)
    total_paid = sum(Decimal(str(r["Payment"])) for r in rows)
    print("\n--- Summary ---")
    print(f"Payments made: {len(rows)}")
    print(f"Total interest: ${money(total_interest)}")
    print(f"Total paid:     ${money(total_paid)}")

    # Show first few lines as a sanity check
    show_n = 5
    print(f"\nFirst {show_n} rows:")
    for r in rows[:show_n]:
        print(r)


if __name__ == "__main__":
    main()
